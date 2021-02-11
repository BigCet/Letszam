import os,json
from openpyxl import Workbook, load_workbook

DATA_FILE = "muszak.xlsx"
# TULORA_FILE = "tuloracsoportos.xlsx"


wb = Workbook()
wbl = load_workbook(filename = DATA_FILE)
ws = wb.active
tulora = load_workbook("tuloracsoportos.xlsx")
lapp = tulora.active
# ws["A1"] = "TSZ"
# ws["B1"] = "Név"
# ws["C1"] = "Terület"
# ws["D1"] = "Beosztás"

sheet_ranges = wbl["2021 Január"]
# print(sheet_ranges["A8"].value)
a = 4
b = 11
for row in range(a, 67):
    a = a+1
    sheet_ranges = wbl["2021 Január"]
    if not (sheet_ranges[f"C{a}"].value):continue

    elif not (sheet_ranges[f"F{a}"].value !="P"): continue
    b = b+1
    # print(b, a)
    print(sheet_ranges[f"B{a}"].value, sheet_ranges[f"C{a}"].value)
    # print(sheet_ranges[f"B{a}"].value, sheet_ranges[f"C{a}"].value)
    # print(sheet_ranges[f"B{a}"].value)
    tsz = ws[f"A{b}"] = sheet_ranges[f"B{a}"].value
    name = ws[f"B{b}"] = sheet_ranges[f"C{a}"].value

    uj_adat = [tsz, name]

    for row in lapp[f"b{b}": f"c{b}"]:
        for index, cell in enumerate(row):
            cell.value = uj_adat[index]

#    tulora.save("tulora_aktualis.xlsx")

# tulora = load_workbook("tuloracsoportos.xlsx")
# lapp = tulora.active

# uj_adat = [tsz, name]
#
# for row in lapp[f"b{b}": f"c{b}"]:
#     for index, cell in enumerate(row):
#         cell.value = uj_adat[index]

tulora.save("tulora_aktualis.xlsx")