import os,json, openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


# DATA_FILE = "Műszaknapló_H4_2021.xlsx"
DATA_FILE = input("Kérem az excel file nevét amiből dolgozzak:")
WORK_FILE = input("Kérem az excel file nevét amibe dolgozzak:")
nap = int(input("Hanyadikára?:"))
muszak = int(input("Napszak?(1, 2, vagy 3):"))

if DATA_FILE == "":
    DATA_FILE = "Műszaknapló_H4_2021.xlsx"

if WORK_FILE == "":
    WORK_FILE = "tuloracsoportos.xlsx"



tol = 6
ig = 14
ora = 8
erend_cod = 2
erend_cod2 = 1

if muszak == 1:
    tol = 6
    ig = 14
elif muszak == 2:
    tol = 14
    ig = 22
elif muszak == 3:
    tol = 22
    ig = 6


wb = Workbook()
wbl = load_workbook(filename = DATA_FILE)
ws = wb.active
tulora = load_workbook(filename = WORK_FILE)
lapp = tulora.active

lapp["C3"] = "2021. Február"

oszlop = nap + 4
a = 3
b = 10
c = -1

sorszam = 0
for row in range(a, 59):
    a = a+1
    sheet_ranges = wbl["2021 Február"]
    if not (sheet_ranges[f"C{a}"].value):continue

    elif sheet_ranges.cell(row=a, column=oszlop).value != muszak : continue
    # elif (sheet_ranges[f"Q{a}"].value != muszak): continue
    b = b+1
    c = c+1
    # print(b, a)
    print(sheet_ranges[f"B{a}"].value, sheet_ranges[f"C{a}"].value, nap)

    tsz = ws[f"B{b}"] = sheet_ranges[f"B{a}"].value
    name = ws[f"C{b}"] = sheet_ranges[f"C{a}"].value


    uj_adat = [tsz, name, nap, tol, ig, 8, 2, "", "", "", "", 1]

    if c == 25:
        png_loc = "matra.jpg"
        my_png = openpyxl.drawing.image.Image(png_loc)
        lapp.add_image(my_png, 'A1')

        tulora.save(f"tulora_aktualis{sorszam}.xlsx")

        sorszam = sorszam+1
        c = 0
        b = 11

    for row in lapp[f"b{b}": f"m{b}"]:
        for index, cell in enumerate(row):
            cell.value = uj_adat[index]
            print(c, b, )

for row in range(b, 35):
    b = b+1

    tsz = ws[f"B{b}"] = ""
    name = ws[f"C{b}"] = ""
    nap = ws[f"D{b}"] = ""
    tol = ws[f"E{b}"] = ""
    ig = ws[f"F{b}"] = ""
    ora = ws[f"G{b}"] = ""
    erend_cod = ws[f"H{b}"] = ""
    erend_cod2 = ws[f"M{b}"] = ""


    uj_adat = [tsz, name, nap, tol, ig, ora, erend_cod, "", "", "", "", erend_cod2]


    for row in lapp[f"b{b}": f"m{b}"]:
        for index, cell in enumerate(row):
            cell.value = uj_adat[index]
            print(c, b)

# wb = load_workbook('test.xlsx')
# ws = wb.active
# png_loc = "matra.jpg"
#
# my_png = openpyxl.drawing.image.Image(png_loc)
# lapp.add_image(my_png, 'A1')
# wb.save('test.xlsx')

tulora.save(f"tulora_aktualis{sorszam+1}.xlsx")


