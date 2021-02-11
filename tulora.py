import os,json
from openpyxl import Workbook, load_workbook

DATA_FILE = "muszak.xlsx"
TULORA_FILE = "tuloracsoportos.xlsx"


wb = Workbook()
wbl = load_workbook(filename = DATA_FILE)
ws = wb.active
ws.cell(row=1, column=1).value = "TSZ"
ws.cell(row=1, column=3).value = "Név"
# ws["C1"] = "Terület"
# ws["D1"] = "Beosztás"

sheet_ranges = wbl["2021 Január"]
# print(sheet_ranges["A8"].value)
a = 4
b = 1
for row in range(a, 67):
    a = a+1
    sheet_ranges = wbl["2021 Január"]
    if not (sheet_ranges[f"C{a}"].value):continue

    elif not (sheet_ranges[f"F{a}"].value !="P"): continue
    b = b+1
    # print(b, a)
    print(sheet_ranges[f"B{a}"].value, sheet_ranges[f"C{a}"].value)
    # print(sheet_ranges[f"B{a}"].value, sheet_ranges[f"C{a}"].value)
    # print(sheet_ranges[f"B{a}"].value)
    ws[f"A{b}"] = sheet_ranges[f"B{a}"].value
    ws[f"B{b}"] = sheet_ranges[f"C{a}"].value

excel_file = os.path.join("tuloracsoportos1.xlsx")
wb.save(excel_file)
